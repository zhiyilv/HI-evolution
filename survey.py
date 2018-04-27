from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException as cannotfind
from selenium.webdriver.common.by import By
import time
import openpyxl
import os
import pickle


file_name = 'HI survey_Springer'


def search_springer(keys_all_words='', keys_exact_phrase='', keys_least_words=[],
                    keys_without_words=[], keys_title=[]):
    # load the advanced search page
    mybrowser = webdriver.Chrome()
    springer_advanced = 'https://link.springer.com/advanced-search'
    mybrowser.get(springer_advanced)

    # find blanks for keywords
    blank_all_words = mybrowser.find_element_by_css_selector('input.all-words')
    blank_exact_phrase = mybrowser.find_element_by_css_selector('input.exact-phrase')
    blank_least_words = mybrowser.find_element_by_css_selector('input.least-words')
    blank_without_words = mybrowser.find_element_by_css_selector('input.without-words')
    blank_title_is = mybrowser.find_element_by_css_selector('input.title-is')

    # send keywords
    blank_all_words.send_keys(keys_all_words)
    blank_exact_phrase.send_keys(keys_exact_phrase)
    blank_least_words.send_keys(' '.join(keys_least_words))
    blank_without_words.send_keys(' '.join(keys_without_words))
    blank_title_is.send_keys(' '.join(keys_title))
    # time.sleep(2)

    # search and return
    button_cookie = mybrowser.find_element_by_css_selector('button.cookie-banner__right-ui')
    button_cookie.click()
    button_submit = mybrowser.find_element_by_css_selector('button#submit-advanced-search.btn.btn-primary.btn-monster')
    button_submit.click()
    time.sleep(3)

    # restrain to articles
    # button_filter = mybrowser.find_element_by_css_selector('a.facet-link')  # the first one is article
    # button_filter.click()
    for bf in mybrowser.find_elements_by_css_selector('a.facet-link'):
        if 'Article' in bf.text:
            bf.click()
            break
    time.sleep(2)
    return mybrowser


def parsing_logging(browser, by, sel_keys, multiple=False):
    try:
        if multiple:
            return browser.find_elements(by, sel_keys)
        else:
            return browser.find_element(by, sel_keys).text.strip()
    except Exception as e:
        with open('log.log', 'a') as f:
            f.write('{}, error:{}, time:{}\n'.format(browser.current_url, e, time.asctime()))
        return None


def parse_search_page(b):
    return [[art.get_attribute('text').strip(), art.get_attribute('href')]
            for art in parsing_logging(b, By.CSS_SELECTOR, 'a.title', multiple=True)]


def parse_search_results(b):
    print('processing search results...')
    search_results = []
    while True:
        title_links = parse_search_page(b)
        for tl in title_links:
            search_results.append([*tl, *([''] * 9)])
        print('added {} papers into results...'.format(len(title_links)))
        try:
            b.find_element_by_css_selector('a.next').click()
            time.sleep(2)
        except cannotfind:
            break
    print('totally {} papers are added.\n'.format(len(search_results)))
    return search_results


def visit_article(url):
    browser_article = webdriver.Chrome()
    browser_article.get(url)
    time.sleep(2)
    try:
        browser_article.find_element_by_id('AboutThisContent').click()  # doi
        browser_article.find_element_by_id('Bib1').click()  # references
    except Exception as e:
        with open('log.log', 'a') as log_f:
            log_f.write('{}, cannot click references or doi:{}, time:{}\n'.format(browser_article.current_url, e, time.asctime()))

    # journal_name = browser_article.find_element_by_css_selector('span.JournalTitle').text
    # type_article = browser_article.find_element_by_css_selector('span.test-render-category').text
    # date = browser_article.find_element_by_css_selector('a.gtm-first-online').text
    # abstract = browser_article.find_element_by_css_selector('p.Para').text
    # doi = browser_article.find_element_by_id('doi-url').text

    journal_name = parsing_logging(browser_article, By.CSS_SELECTOR, 'span.JournalTitle') or ''
    type_article = parsing_logging(browser_article, By.CSS_SELECTOR, 'span.test-render-category') or 'article'
    date = parsing_logging(browser_article, By.CLASS_NAME, 'article-dates__entry') or ''
    abstract = parsing_logging(browser_article, By.CSS_SELECTOR, 'p.Para') or ''
    doi = parsing_logging(browser_article, By.ID, 'doi-url') or ''
    citation_count = parsing_logging(browser_article, By.ID, 'citations-count-number') or 0
    citation_count = int(citation_count)

    author_list = parsing_logging(browser_article, By.CSS_SELECTOR, 'span.authors__name', multiple=True) or []
    author_list = [au.text.strip() for au in author_list]

    keyword_list = parsing_logging(browser_article, By.CSS_SELECTOR, 'span.Keyword', multiple=True) or []
    keyword_list = [k.text.strip() for k in keyword_list]

    reference_list = parsing_logging(browser_article, By.CLASS_NAME, 'CitationContent', multiple=True) or []
    reference_list = [r.text.strip() for r in reference_list]

    browser_article.close()
    return journal_name, author_list, type_article, date, citation_count, abstract, keyword_list, reference_list, doi


def update_pickle(fn, line_number):
    with open('{}.pickle'.format(fn), 'rb') as f:
        data_pickle = pickle.load(f)
    paper_info = data_pickle[line_number]
    print('processing #{}: {}...'.format(line_number, paper_info[0]), end=' ')
    paper_info[2:] = visit_article(paper_info[1])  # get detailed info for the paper

    # store into pickle
    data_pickle[line_number] = paper_info
    with open('{}.pickle'.format(fn), 'wb') as f:
        pickle.dump(data_pickle, f)
    print('written into pickle.')

    # store into excel
    # piece = [title, link, *info]
    # piece = ['\n'.join(i) if isinstance(i, list) else i for i in piece]
    # data_excel = openpyxl.load_workbook('{}.xlsx'.format(fn))
    # data_excel.active.append(piece)
    # data_excel.save('{}.xlsx'.format(fn))
    # print('written into excel...')


def pickle2excel(fn):
    with open('{}.pickle'.format(fn), 'rb') as f:
        data_pickle = pickle.load(f)
    if '{}.xlsx'.format(fn) not in os.listdir(os.getcwd()):
        wb = openpyxl.Workbook()
        for line in data_pickle:
            line = ['\n'.join(i) if isinstance(i, list) else i for i in line]
            wb.active.append(line)
        wb.save('{}.xlsx'.format(fn))


if __name__ == '__main__':
    input_title = ['happiness']
    input_at_least_one = ['air', 'environment', 'hazards', 'pollution']
    input_all_words = 'happiness scale measurement'
    search_result_page = search_springer(keys_all_words=input_all_words,
                                         keys_least_words=input_at_least_one,
                                         keys_title=input_title)

    header = ['title', 'link', 'journal_name', 'author_list', 'type', 'date', 'citation_count', 'abstract',
              'keywords', 'references', 'doi']

    sr = parse_search_results(search_result_page)
    results = [header, *sr]

    # initialization for excel and pickle file
    # if '{}.xlsx'.format(file_name) not in os.listdir(os.getcwd()):
    #     wb = openpyxl.Workbook()
    #     for line in results:
    #         wb.active.append(line)
    #     wb.save('{}.xlsx'.format(file_name))
    if '{}.pickle'.format(file_name) not in os.listdir(os.getcwd()):
        with open('{}.pickle'.format(file_name), 'wb') as p_ref:
            pickle.dump(results, p_ref)
    print('saved in excel and pickle files.\n')

    # process each paper
    start = 1  # the starting number of paper
    for index, paper in enumerate(results[start:]):
        update_pickle(fn=file_name, line_number=index+start)

    # store into excel for good readings
    pickle2excel(file_name)










